import openpyxl
from datetime import datetime, timedelta
import os
import requests

# === CONFIG ===
MENU_PATH = "menu.xlsx"
DAYSTART_PATH = "day_to_start.txt"
TELEGRAM_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]

from datetime import date

# === DATE DA ESCLUDERE (MODIFICA QUI) ===
# - Giorni singoli: "YYYY-MM-DD"
# - Range inclusivi: ("YYYY-MM-DD", "YYYY-MM-DD")
# - Range aperti: (None, "YYYY-MM-DD") oppure ("YYYY-MM-DD", None)
EXCLUDED_DATES = [
    #"2025-12-24",
     ("2025-12-24", "2026-01-06")
    # (None, "2026-01-06"),
    # ("2025-12-27", None),
]

def _parse_iso_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()

def is_excluded(d: date) -> bool:
    for item in EXCLUDED_DATES:
        # giorno singolo
        if isinstance(item, str):
            if _parse_iso_date(item) == d:
                return True
        # range
        else:
            start_s, end_s = item
            start = _parse_iso_date(start_s) if start_s else date.min
            end = _parse_iso_date(end_s) if end_s else date.max
            if start <= d <= end:
                return True
    return False


def giorni_lavorativi_da_a(data_inizio, data_fine):
    giorni = 0
    giorno = data_inizio
    while giorno < data_fine:
        #if giorno.weekday() < 5:  # 0=lun, 4=ven
        if giorno.weekday() < 5 and not is_excluded(giorno):
            giorni += 1
        giorno += timedelta(days=1)
    return giorni

def componi_messaggio_menu(menu_del_giorno, giorno_settimana, data_it):
    msg = (
        f"Buongiorno e buon lavoro.\n\n"
        f"🧑‍🍳 *Menù del giorno* ({giorno_settimana.title()} {data_it})\n\n"
        f"*[Primi]*\n"
        f"{menu_del_giorno[0]}.\n"
        f"{menu_del_giorno[1]}.\n"
        f"{menu_del_giorno[2]}.\n"
        f"*[Pasta o riso in bianco/pomodoro]*\n"
        f"*[Secondi]*\n"
        f"{menu_del_giorno[3]}.\n"
        f"{menu_del_giorno[4]}.\n"
        f"{menu_del_giorno[5]}.\n"
        f"*[Pizza gusti del giorno]*\n"
        f"*[Contorni]*\n"
        f"{menu_del_giorno[6]}.\n"       
        f"\nBuon appetito dalla Commissione mensa.\n👋"
    )
    return msg

def parse_giorno_settimana(s):
    giorni_sett = ["LUNEDI", "MARTEDÌ", "MERCOLEDÌ", "GIOVEDÌ", "VENERDÌ"]
    for g in giorni_sett:
        if s.startswith(g):
            n = int(s.replace(g, "").strip())
            return g, n
    raise ValueError("Formato giorno_settimana errato")

def trova_riga_col_settimane(ws):
    for row in ws.iter_rows(min_row=1, max_row=15):
        valori = [str(cell.value).upper() if cell.value else "" for cell in row]
        if any("SETTIMANA" in v for v in valori):
            return row, valori
    raise ValueError("Non trovata riga settimane!")

def trova_blocchi_giorni(ws):
    giorni = ["LUNEDI", "MARTEDÌ", "MERCOLEDÌ", "GIOVEDÌ", "VENERDÌ"]
    blocchi = []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        prima_col = str(row[0]).upper() if row[0] else ""
        if prima_col in giorni:
            blocchi.append( (prima_col, i+1) )
    return blocchi

def trova_colonna_settimana(intestazioni, settimana_n):
    for idx, v in enumerate(intestazioni):
        if v.strip().upper() == f"SETTIMANA {settimana_n}":
            return idx
    raise ValueError("Settimana non trovata")

def trova_blocco_per_giorno(blocchi, giorno):
    for nome, riga in blocchi:
        if nome == giorno:
            return riga
    raise ValueError("Giorno non trovato")

def estrai_menu(ws, riga_giorno, col_settimana):
    menu = []
    NUM_VOCI_MENU = 7
    for r in range(riga_giorno, riga_giorno + NUM_VOCI_MENU):
        val = ws.cell(row=r, column=col_settimana+1).value
        if val: menu.append(str(val))
    return menu

# === LEGGI DAY_TO_START ===
with open(DAYSTART_PATH, encoding="utf-8") as f:
    daystart = f.read().strip()
giorno_start, data_start = [x.strip() for x in daystart.split(",")]

d_start = datetime.strptime(data_start, "%Y-%m-%d").date()
d_oggi = datetime.today().date()   # OGGI

# Salta se la data è esclusa
if is_excluded(d_oggi):
    print(f"Oggi {d_oggi} è escluso (EXCLUDED_DATES). Nessun menu da pubblicare.")
    exit(0)

# Salta il weekend (solo pubblicazione giorni lavorativi)
if d_oggi.weekday() >= 5:
    print("Oggi è sabato/domenica, nessun menu da pubblicare.")
    exit(0)

giorni_sett = ["LUNEDI", "MARTEDÌ", "MERCOLEDÌ", "GIOVEDÌ", "VENERDÌ"]
g_start, sett_start = parse_giorno_settimana(giorno_start)
idx_giorno_start = giorni_sett.index(g_start)
delta_days = giorni_lavorativi_da_a(d_start, d_oggi)
pos_start = (sett_start - 1) * 5 + idx_giorno_start
pos_oggi = pos_start + delta_days
settimane_totali = 4
settimana_menu = (pos_oggi // 5) % settimane_totali + 1
giorno_menu = giorni_sett[pos_oggi % 5]

# --- Estrai menu
wb = openpyxl.load_workbook(MENU_PATH, data_only=True)
ws = wb.worksheets[0]  # Primo foglio del file Excel
row_sett, intestazioni = trova_riga_col_settimane(ws)
blocchi = trova_blocchi_giorni(ws)
riga_giorno = trova_blocco_per_giorno(blocchi, giorno_menu)
col_settimana = trova_colonna_settimana(intestazioni, settimana_menu)
menu = estrai_menu(ws, riga_giorno, col_settimana)

# --- Componi messaggio
data_it = d_oggi.strftime("%d/%m/%Y")
msg = componi_messaggio_menu(menu, giorno_menu, data_it)

# --- Manda su telegram
def send_telegram_message(token, chat_id, text):
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "Markdown"
    }
    r = requests.post(url, json=payload)
    print(r.text)

send_telegram_message(
    TELEGRAM_TOKEN,
    TELEGRAM_CHAT_ID,
    msg
)
